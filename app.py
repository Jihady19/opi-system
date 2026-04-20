"""
OPI - Oil Production Improvement System
Neural Network based optimization and forecasting for oil wells
"""

import os
import pickle
import sqlite3
import numpy as np
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, render_template, g, send_file
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler
from scipy.optimize import minimize
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import Pipeline
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error
import io
import threading
import tempfile
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

app = Flask(__name__)

# ── Cross-platform base directory (same folder as app.py) ──
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app.config['DATABASE']         = os.path.join(BASE_DIR, 'opi_data.db')
app.config['MODEL_PATH']       = os.path.join(BASE_DIR, 'model.pkl')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['UPLOAD_FOLDER']    = os.path.join(BASE_DIR, 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ------------------- Database -------------------
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    db = sqlite3.connect(app.config['DATABASE'])
    db.execute("""
        CREATE TABLE IF NOT EXISTS production_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            total_liquid REAL,
            oil_production REAL,
            water_production REAL,
            water_cut REAL,
            upstream_pres REAL,
            downstream_pres REAL,
            choke_size REAL,
            gor REAL,
            file_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS uploaded_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            records_added INTEGER,
            date_range_start TEXT,
            date_range_end TEXT,
            uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    try:
        db.execute("ALTER TABLE production_data ADD COLUMN gor REAL")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE production_data ADD COLUMN file_id INTEGER")
    except sqlite3.OperationalError:
        pass
    db.commit()
    db.close()

# ------------------- Neural Network Model (Optimization) -------------------
class OilProductionModel:
    def __init__(self):
        self.model = None
        self.scaler_X = StandardScaler()
        self.scaler_y = StandardScaler()
        self.is_trained = False
        self.train_samples = 0
        self.score = 0.0
        self.last_trained = None
        self.epochs = 200        # ← تقليل من 500 إلى 200 للسرعة
        self.batch_size = 32
        self.bounds = {
            'upstream_pres':   (0, 58),
            'downstream_pres': (0, 38.6),
            'choke_size':      (0, 128)
        }

    def _build_model(self):
        return MLPRegressor(
            hidden_layer_sizes=(128, 64, 32),
            activation='relu',
            solver='adam',
            max_iter=self.epochs,
            batch_size=self.batch_size,
            random_state=42,
            early_stopping=True,
            validation_fraction=0.1,
            n_iter_no_change=10,   # ← من 20 إلى 10 للتوقف المبكر
            learning_rate_init=0.001,
            warm_start=False
        )

    def train(self, df: pd.DataFrame):
        required_cols = ['total_liquid', 'oil_production', 'water_production',
                         'water_cut', 'upstream_pres', 'downstream_pres', 'choke_size']
        df = df.dropna(subset=required_cols)
        if len(df) < 10:
            raise ValueError("Insufficient data for training (less than 10 records)")
        X = df[required_cols].values.astype(float)
        y = df[['oil_production', 'water_production', 'water_cut',
                'upstream_pres', 'downstream_pres']].values.astype(float)
        X_scaled = self.scaler_X.fit_transform(X)
        y_scaled = self.scaler_y.fit_transform(y)
        self.model = self._build_model()
        self.model.fit(X_scaled, y_scaled)
        y_pred = self.scaler_y.inverse_transform(self.model.predict(X_scaled))
        ss_res = np.sum((y - y_pred) ** 2)
        ss_tot = np.sum((y - np.mean(y, axis=0)) ** 2)
        self.score = float(1 - ss_res / ss_tot) if ss_tot != 0 else 0.0
        self.is_trained = True
        self.train_samples = len(df)
        self.last_trained = datetime.now().strftime('%Y-%m-%d %H:%M')

    def predict(self, features: dict) -> dict:
        if not self.is_trained:
            raise RuntimeError("Model not trained")
        cols = ['total_liquid', 'oil_production', 'water_production',
                'water_cut', 'upstream_pres', 'downstream_pres', 'choke_size']
        X = np.array([[features.get(c, 0) or 0 for c in cols]], dtype=float)
        X_scaled = self.scaler_X.transform(X)
        y_scaled = self.model.predict(X_scaled)
        y = self.scaler_y.inverse_transform(y_scaled)[0]
        return {
            'oil_production':   float(y[0]),
            'water_production': float(y[1]),
            'water_cut':        float(np.clip(y[2], 0, 100)),
            'upstream_pres':    float(np.clip(y[3], *self.bounds['upstream_pres'])),
            'downstream_pres':  float(np.clip(y[4], *self.bounds['downstream_pres']))
        }

    def optimize(self, features: dict) -> dict:
        if not self.is_trained:
            raise RuntimeError("Model not trained")
        cols = ['total_liquid', 'oil_production', 'water_production',
                'water_cut', 'upstream_pres', 'downstream_pres', 'choke_size']
        def objective(params):
            upstream, downstream, choke = params
            f = dict(features)
            f['upstream_pres']   = upstream
            f['downstream_pres'] = downstream
            f['choke_size']      = choke
            X = np.array([[f.get(c, 0) or 0 for c in cols]], dtype=float)
            X_sc = self.scaler_X.transform(X)
            y_sc = self.model.predict(X_sc)
            y = self.scaler_y.inverse_transform(y_sc)[0]
            oil = y[0]
            wc  = np.clip(y[2], 0, 100)
            return -oil + 0.5 * wc
        x0 = [
            features.get('upstream_pres',   30) or 30,
            features.get('downstream_pres', 20) or 20,
            features.get('choke_size',       32) or 32
        ]
        bounds = [self.bounds['upstream_pres'], self.bounds['downstream_pres'], self.bounds['choke_size']]
        result = minimize(objective, x0, method='L-BFGS-B', bounds=bounds, options={'maxiter': 1000, 'ftol': 1e-9})
        opt_up, opt_down, opt_choke = result.x
        f_opt = dict(features)
        f_opt['upstream_pres']   = opt_up
        f_opt['downstream_pres'] = opt_down
        f_opt['choke_size']      = opt_choke
        pred = self.predict(f_opt)
        return {
            'upstream_pres':   float(np.clip(opt_up,    *self.bounds['upstream_pres'])),
            'downstream_pres': float(np.clip(opt_down,  *self.bounds['downstream_pres'])),
            'choke_size':      float(np.clip(opt_choke, *self.bounds['choke_size'])),
            'predicted_oil':        pred['oil_production'],
            'predicted_water':      pred['water_production'],
            'predicted_water_cut':  pred['water_cut'],
        }

model_instance = OilProductionModel()

def save_model():
    with open(app.config['MODEL_PATH'], 'wb') as f:
        pickle.dump(model_instance, f)

def load_model():
    global model_instance
    if os.path.exists(app.config['MODEL_PATH']):
        with open(app.config['MODEL_PATH'], 'rb') as f:
            model_instance = pickle.load(f)

# ------------------- Data Helpers -------------------
COL_MAP = {
    # Arabic names
    'تاريخ': 'date', 'السائل المستخرج': 'total_liquid', 'النفط المستخرج': 'oil_production',
    'الماء المستخرج': 'water_production', 'نسبة الماء': 'water_cut', 'نسبة الماء ': 'water_cut',
    'الضغط الأعلى': 'upstream_pres', 'الضغط الأدنى': 'downstream_pres', 'الضغط الأدنى ': 'downstream_pres',
    'فتحة الخانق': 'choke_size',
    # English spaced format (old Excel)
    'THEDATE': 'date', 'DAILY LIQUID': 'total_liquid', 'DAILYOIL': 'oil_production',
    'QW': 'water_production', 'WATER CUT': 'water_cut',
    'UPSTREAMPRES': 'upstream_pres', 'DOWNSTREAMPRES': 'downstream_pres', 'CHOKESIZE': 'choke_size',
    # English compact format (AG-22 style)
    'DAILYLIQUID': 'total_liquid', 'WATERCUT': 'water_cut',
    'CHOKE_SIZE': 'choke_size', 'UPSTREAM_PRES': 'upstream_pres', 'DOWNSTREAM_PRES': 'downstream_pres',
    'UPSTREAM': 'upstream_pres', 'DOWNSTREAM': 'downstream_pres',
    'DATE': 'date', 'OIL': 'oil_production', 'WATER': 'water_production', 'LIQUID': 'total_liquid',
    # GOR variants
    'GOR': 'gor', 'GOR_VERIFY': 'gor', 'gor': 'gor',
}

def parse_excel(filepath) -> pd.DataFrame:
    """
    Parse Excel files with single or double header rows, and any column naming style.
    Supports: Arabic names, spaced English (DAILY LIQUID), compact English (DAILYLIQUID / AG-22 style).
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        raise ValueError("Excel file is empty")

    row1 = all_rows[0]
    row2 = all_rows[1] if len(all_rows) > 1 else None

    def is_header_row(row):
        if row is None:
            return False
        non_none = [v for v in row if v is not None]
        if not non_none:
            return False
        return sum(1 for v in non_none if isinstance(v, str)) >= len(non_none) * 0.7

    if row2 is not None and is_header_row(row2):
        # Double-header: row1=Arabic, row2=English — prefer whichever maps to a known column
        merged_headers = []
        for h1, h2 in zip(row1, row2):
            h1s = str(h1).strip() if h1 is not None else ''
            h2s = str(h2).strip() if h2 is not None else ''
            if h2s and h2s in COL_MAP:
                merged_headers.append(h2s)
            elif h1s and h1s in COL_MAP:
                merged_headers.append(h1s)
            elif h2s:
                merged_headers.append(h2s)
            else:
                merged_headers.append(h1s)
        data_rows = all_rows[2:]
    else:
        merged_headers = [str(h).strip() if h is not None else '' for h in row1]
        data_rows = all_rows[1:]

    data = [dict(zip(merged_headers, row)) for row in data_rows if any(c is not None for c in row)]
    df = pd.DataFrame(data)

    if df.empty:
        try:
            df = pd.read_excel(filepath, engine='openpyxl', header=0)
        except Exception:
            df = pd.read_excel(filepath, engine='openpyxl')

    # Rename columns using COL_MAP (strip whitespace first)
    df.rename(columns=lambda c: COL_MAP.get(str(c).strip(), str(c).strip()), inplace=True)

    # Drop serial-number / unnamed columns (e.g. 'ت')
    drop_cols = [c for c in df.columns if c in ('ت', '', 'None') or
                 (len(str(c)) <= 2 and not str(c).isalpha())]
    df.drop(columns=drop_cols, errors='ignore', inplace=True)

    # Remove duplicate mapped columns — keep first occurrence only
    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    # Convert numeric columns
    num_cols = ['total_liquid', 'oil_production', 'water_production', 'water_cut',
                'upstream_pres', 'downstream_pres', 'choke_size', 'gor']
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Parse date column
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.strftime('%Y-%m-%d')

    return df

def db_to_dataframe() -> pd.DataFrame:
    conn = sqlite3.connect(app.config['DATABASE'])
    df = pd.read_sql("SELECT * FROM production_data", conn)
    conn.close()
    return df

def get_all_data():
    """إرجاع جميع سجلات الإنتاج كقائمة من القواميس مع ضمان القيم الرقمية"""
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute("SELECT * FROM production_data ORDER BY date").fetchall()
        result = []
        for r in rows:
            d = dict(r)
            # ضمان أن القيم الرقمية ليست None
            for col in ['total_liquid', 'oil_production', 'water_production', 'water_cut']:
                d[col] = float(d[col]) if d[col] is not None else 0.0
            d['gor'] = float(d['gor']) if d.get('gor') is not None else None
            result.append(d)
        return result
    finally:
        conn.close()

# ------------------- Page Routes -------------------
@app.route('/')
def index(): return render_template('index.html')
@app.route('/optimize')
def optimize_page(): return render_template('optimize.html')
@app.route('/data')
def data_page(): return render_template('data.html')
@app.route('/upload')
def upload_page(): return render_template('upload.html')
@app.route('/model')
def model_page(): return render_template('model.html')
@app.route('/forecast')
def forecast_page(): return render_template('forecast.html')

# ------------------- API Routes (existing) -------------------
@app.route('/api/stats')
def api_stats():
    db = get_db()
    row = db.execute("SELECT COUNT(*) as cnt, AVG(water_cut) as avg_wc, AVG(oil_production) as avg_oil FROM production_data").fetchone()
    return jsonify({
        'total_records': row['cnt'],
        'avg_water_cut': row['avg_wc'] or 0,
        'avg_oil': row['avg_oil'] or 0,
        'model_ready': model_instance.is_trained,
        'model_training': not model_instance.is_trained  # ← مفيد للـ frontend يعرض "جاري التدريب..."
    })

@app.route('/api/data')
def api_data():
    db = get_db()
    rows = db.execute("SELECT * FROM production_data ORDER BY date DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/data/<int:row_id>', methods=['DELETE'])
def api_delete_row(row_id):
    db = get_db()
    db.execute("DELETE FROM production_data WHERE id=?", (row_id,))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/data/all', methods=['DELETE'])
def api_delete_all():
    db = get_db()
    db.execute("DELETE FROM production_data")
    db.commit()
    return jsonify({'success': True})

@app.route('/api/years')
def api_years():
    db = get_db()
    rows = db.execute("SELECT strftime('%Y', date) as year, COUNT(*) as count FROM production_data WHERE date IS NOT NULL AND date != '' GROUP BY year ORDER BY year DESC").fetchall()
    return jsonify([{'year': int(r['year']), 'count': r['count']} for r in rows])

@app.route('/api/data/year/<int:year>')
def api_data_by_year(year):
    db = get_db()
    rows = db.execute("SELECT * FROM production_data WHERE strftime('%Y', date) = ? ORDER BY date", (str(year),)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/download/year/<int:year>')
def download_year_excel(year):
    db = get_db()
    rows = db.execute("SELECT date, total_liquid, oil_production, water_production, water_cut, upstream_pres, downstream_pres, choke_size, gor FROM production_data WHERE strftime('%Y', date) = ? ORDER BY date", (str(year),)).fetchall()
    if not rows:
        return jsonify({'error': 'No data for this year'}), 404
    df = pd.DataFrame([dict(r) for r in rows])
    tmp_dir = tempfile.gettempdir()
    tmp = os.path.join(tmp_dir, f'production_{year}.xlsx')
    df.to_excel(tmp, index=False)
    return send_file(tmp, as_attachment=True, download_name=f'production_{year}.xlsx')

@app.route('/api/data/timeseries')
def api_timeseries():
    years = request.args.get('years', '')
    if not years:
        return jsonify([])
    year_list = [int(y.strip()) for y in years.split(',') if y.strip().isdigit()]
    if not year_list:
        return jsonify([])
    db = get_db()
    placeholders = ','.join('?' * len(year_list))
    rows = db.execute(f"SELECT date, oil_production, water_production, water_cut, gor FROM production_data WHERE strftime('%Y', date) IN ({placeholders}) ORDER BY date", [str(y) for y in year_list]).fetchall()
    data = [{'date': r['date'], 'oil_production': r['oil_production'], 'water_production': r['water_production'], 'water_cut': r['water_cut'], 'gor': r['gor']} for r in rows]
    df = pd.DataFrame(data)
    if df.empty:
        return jsonify([])
    df['date'] = pd.to_datetime(df['date'])
    df['year'] = df['date'].dt.year
    df['month'] = df['date'].dt.month
    monthly_oil = df.groupby(['year','month'])['oil_production'].mean().reset_index()
    monthly_water = df.groupby(['year','month'])['water_production'].mean().reset_index()
    monthly_wcut = df.groupby(['year','month'])['water_cut'].mean().reset_index()
    monthly_gor = df.groupby(['year','month'])['gor'].mean().reset_index()
    result = []
    for _, row in monthly_oil.iterrows():
        y = int(row['year']); m = int(row['month'])
        oil = row['oil_production']
        water = monthly_water[(monthly_water['year']==y)&(monthly_water['month']==m)]['water_production'].values
        wcut = monthly_wcut[(monthly_wcut['year']==y)&(monthly_wcut['month']==m)]['water_cut'].values
        gor = monthly_gor[(monthly_gor['year']==y)&(monthly_gor['month']==m)]['gor'].values
        result.append({
            'year': y, 'month': m,
            'oil_production': round(oil,2),
            'water_production': round(water[0],2) if len(water)>0 else 0,
            'water_cut': round(wcut[0],2) if len(wcut)>0 else 0,
            'gor': round(gor[0],2) if len(gor)>0 else 0
        })
    return jsonify(result)

@app.route('/api/upload', methods=['POST'])
def api_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    original_filename = file.filename
    if not original_filename.endswith(('.xlsx','.xls')):
        return jsonify({'error': 'Unsupported file format. Please upload Excel file'}), 400
    tmp = os.path.join(tempfile.gettempdir(), f'upload_{datetime.now().timestamp()}.xlsx')
    file.save(tmp)
    try:
        df = parse_excel(tmp)
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    if 'oil_production' not in df.columns or 'water_cut' not in df.columns:
        return jsonify({'error': 'Required columns not found: oil_production, water_cut'}), 400
    db = get_db()
    # Register the file in uploaded_files table
    date_start = None
    date_end = None
    if 'date' in df.columns:
        valid_dates = df['date'].dropna()
        if len(valid_dates) > 0:
            date_start = str(valid_dates.min())
            date_end = str(valid_dates.max())
    file_row = db.execute(
        "INSERT INTO uploaded_files (filename, records_added, date_range_start, date_range_end) VALUES (?, 0, ?, ?)",
        (original_filename, date_start, date_end)
    )
    file_id = file_row.lastrowid
    db.commit()
    added = 0
    for _, row in df.iterrows():
        vals = {c: (None if pd.isna(row.get(c, np.nan)) else row.get(c)) for c in
                ['date','total_liquid','oil_production','water_production','water_cut','upstream_pres','downstream_pres','choke_size','gor']}
        if vals['oil_production'] is None:
            continue
        vals['file_id'] = file_id
        db.execute("""INSERT INTO production_data 
            (date, total_liquid, oil_production, water_production, water_cut, upstream_pres, downstream_pres, choke_size, gor, file_id)
            VALUES (:date, :total_liquid, :oil_production, :water_production, :water_cut, :upstream_pres, :downstream_pres, :choke_size, :gor, :file_id)""", vals)
        added += 1
    db.execute("UPDATE uploaded_files SET records_added=? WHERE id=?", (added, file_id))
    db.commit()
    retrain_msg = ''
    try:
        full_df = db_to_dataframe()
        model_instance.train(full_df)
        save_model()
        retrain_msg = f'Model retrained successfully on {len(full_df)} records.'
    except Exception as e:
        retrain_msg = f'Warning: {str(e)}'
    return jsonify({'added_count': added, 'retrain_message': retrain_msg, 'file_id': file_id})

@app.route('/api/files')
def api_files():
    db = get_db()
    rows = db.execute("SELECT * FROM uploaded_files ORDER BY uploaded_at DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/files/<int:file_id>', methods=['DELETE'])
def api_delete_file(file_id):
    db = get_db()
    db.execute("DELETE FROM production_data WHERE file_id=?", (file_id,))
    db.execute("DELETE FROM uploaded_files WHERE id=?", (file_id,))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/files/<int:file_id>/data')
def api_file_data(file_id):
    db = get_db()
    rows = db.execute("SELECT * FROM production_data WHERE file_id=? ORDER BY date", (file_id,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/files/<int:file_id>/timeseries')
def api_file_timeseries(file_id):
    db = get_db()
    rows = db.execute(
        "SELECT date, oil_production, water_production, water_cut, gor, total_liquid FROM production_data WHERE file_id=? ORDER BY date",
        (file_id,)
    ).fetchall()
    data = [dict(r) for r in rows]
    if not data:
        return jsonify([])
    df = pd.DataFrame(data)
    df['date'] = pd.to_datetime(df['date'])
    df['year'] = df['date'].dt.year
    df['month'] = df['date'].dt.month
    result = []
    for (year, month), grp in df.groupby(['year', 'month']):
        result.append({
            'year': int(year), 'month': int(month),
            'oil_production': round(float(grp['oil_production'].mean()), 2),
            'water_production': round(float(grp['water_production'].mean()), 2),
            'water_cut': round(float(grp['water_cut'].mean()), 2),
            'gor': round(float(grp['gor'].mean()), 2) if grp['gor'].notna().any() else 0,
            'total_liquid': round(float(grp['total_liquid'].mean()), 2),
        })
    return jsonify(result)

@app.route('/api/optimize', methods=['POST'])
def api_optimize():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Missing data'}), 400
    if data.get('oil_production') is None or data.get('water_cut') is None:
        return jsonify({'error': 'Oil production and water cut are required'}), 400
    if not model_instance.is_trained:
        try:
            df = db_to_dataframe()
            model_instance.train(df)
            save_model()
        except Exception as e:
            return jsonify({'error': f'Model not ready: {e}'}), 503
    db = get_db()
    avgs = db.execute("SELECT AVG(total_liquid) al, AVG(upstream_pres) up, AVG(downstream_pres) dp, AVG(choke_size) cs FROM production_data").fetchone()
    features = {
        'total_liquid': data.get('total_liquid') or (avgs['al'] or 1200),
        'oil_production': data['oil_production'],
        'water_production': data.get('water_production') or 0,
        'water_cut': data['water_cut'],
        'upstream_pres': data.get('upstream_pres') or (avgs['up'] or 30),
        'downstream_pres': data.get('downstream_pres') or (avgs['dp'] or 21),
        'choke_size': data.get('choke_size') or (avgs['cs'] or 32)
    }
    note = None
    if any(data.get(k) is None for k in ['upstream_pres','downstream_pres','choke_size']):
        note = 'Some operating parameters missing, using historical averages'
    try:
        opt = model_instance.optimize(features)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    oil_gain = opt['predicted_oil'] - features['oil_production']
    wc_reduction = features['water_cut'] - opt['predicted_water_cut']
    confidence = min(0.95, max(0.3, model_instance.score * 0.9 + min(model_instance.train_samples/1000, 0.1)))
    return jsonify({
        'optimized': {
            'upstream_pres': round(opt['upstream_pres'],2),
            'downstream_pres': round(opt['downstream_pres'],2),
            'choke_size': round(opt['choke_size'],1),
            'predicted_oil': round(opt['predicted_oil'],2),
            'predicted_water_cut': round(opt['predicted_water_cut'],2)
        },
        'improvement': {
            'oil_gain': round(max(oil_gain,0),2),
            'water_cut_reduction': round(max(wc_reduction,0),2)
        },
        'confidence': round(confidence,3),
        'note': note
    })

@app.route('/api/model/info')
def api_model_info():
    return jsonify({
        'ready': model_instance.is_trained,
        'train_samples': model_instance.train_samples,
        'score': round(model_instance.score,4) if model_instance.is_trained else None,
        'last_trained': model_instance.last_trained,
        'epochs': model_instance.epochs,
        'batch_size': model_instance.batch_size
    })

@app.route('/api/model/retrain', methods=['POST'])
def api_retrain():
    try:
        df = db_to_dataframe()
        if len(df) < 10:
            return jsonify({'error': 'Insufficient data (less than 10 records)'}), 400
        model_instance.train(df)
        save_model()
        return jsonify({'success': True, 'samples': model_instance.train_samples, 'score': round(model_instance.score,4)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ------------------- Advanced Forecasting Functions -------------------

def build_forecast_monthly(all_data_sorted, years_ahead, rf_model, rf_scaler):
    """
    بناء التنبؤ الشهري بطريقة صحيحة:
    - يبدأ من آخر قيمة تاريخية فعلية (متصل بالخط التاريخي)
    - الاتجاه: متوسط آخر 6 فروق شهرية (مستقر وغير متطرف)
    - التذبذب: نسبي (% من القيمة الحالية) لمنع الصفر
    - موسمية خفيفة من أنماط البيانات التاريخية
    - decay تدريجي للاتجاه مع الزمن
    """
    from datetime import date as date_cls

    # ── بناء DataFrame شهري ──
    df_raw = pd.DataFrame(all_data_sorted)
    df_raw['date'] = pd.to_datetime(df_raw['date'])
    df_raw['ym']   = df_raw['date'].dt.to_period('M')

    monthly = df_raw.groupby('ym').agg(
        oil=('oil_production',    'mean'),
        water=('water_production', 'mean'),
        liq=('total_liquid',      'mean'),
        wc=('water_cut',          'mean'),
    ).reset_index()
    monthly['date'] = monthly['ym'].dt.start_time
    monthly = monthly.sort_values('date').reset_index(drop=True)

    n_hist = len(monthly)
    if n_hist < 3:
        return []

    # ── 1. نقطة البداية = آخر قيمة تاريخية ──
    last_oil   = max(float(monthly['oil'].iloc[-1]),   1.0)
    last_water = max(float(monthly['water'].iloc[-1]), 0.0)
    last_liq   = max(float(monthly['liq'].iloc[-1]),   1.0)
    last_wc    = float(np.clip(monthly['wc'].iloc[-1], 0.0, 100.0))

    # ── 2. الاتجاه: متوسط آخر 6 فروق شهرية ──
    def short_trend(series, n=6):
        """متوسط آخر n فروق شهرية — أكثر استقراراً من Regression"""
        diffs = np.diff(series.values.astype(float))
        recent = diffs[max(-n, -len(diffs)):]
        return float(np.mean(recent)) if len(recent) > 0 else 0.0

    trend_oil   = short_trend(monthly['oil'])
    trend_water = short_trend(monthly['water'])
    trend_liq   = short_trend(monthly['liq'])
    trend_wc    = short_trend(monthly['wc'])

    # ── 3. التذبذب النسبي (% من القيمة) لمنع الصفر ──
    def relative_vol(series, n=12):
        """الانحراف المعياري للتغيرات النسبية الشهرية"""
        vals = series.values.astype(float)
        v_prev = vals[:-1]
        mask = v_prev > 0
        if mask.sum() < 2:
            return 0.05  # 5% افتراضي
        rel = np.diff(vals)[mask] / v_prev[mask]
        return float(np.std(rel[-n:])) if len(rel) >= 2 else 0.05

    vol_oil   = relative_vol(monthly['oil'])
    vol_water = relative_vol(monthly['water'])
    vol_liq   = relative_vol(monthly['liq'])
    vol_wc    = max(relative_vol(monthly['wc']), 0.02)

    # ── 4. الموسمية (seasonal index لكل شهر 1-12) ──
    monthly['month_num'] = monthly['date'].dt.month
    season = monthly.groupby('month_num').agg(
        oil_s=('oil',   'mean'),
        water_s=('water','mean'),
        liq_s=('liq',   'mean'),
        wc_s=('wc',     'mean'),
    )
    for col, mean_val in [('oil_s',   float(monthly['oil'].mean())   or 1),
                           ('water_s', float(monthly['water'].mean()) or 1),
                           ('liq_s',   float(monthly['liq'].mean())   or 1),
                           ('wc_s',    float(monthly['wc'].mean())    or 1)]:
        season[col] = (season[col] / mean_val).clip(0.7, 1.3)
    season = season.fillna(1.0)

    # ── حدود تاريخية ──
    hist_max_oil   = float(monthly['oil'].max())   * 1.3
    hist_max_water = float(monthly['water'].max()) * 1.3
    hist_max_liq   = float(monthly['liq'].max())   * 1.3
    # الحد الأدنى: 5% من آخر قيمة (لمنع الصفر التام)
    floor_oil   = last_oil   * 0.05
    floor_water = last_water * 0.05 if last_water > 1 else 0.0
    floor_liq   = last_liq   * 0.05

    rng = np.random.default_rng(seed=7)

    n_months    = years_ahead * 12
    last_period = monthly['ym'].iloc[-1]
    forecast_list = []

    cur_oil   = last_oil
    cur_water = last_water
    cur_liq   = last_liq
    cur_wc    = last_wc

    for i in range(1, n_months + 1):
        fcast = last_period + i
        m_num = fcast.month

        # معامل الموسمية
        if m_num in season.index:
            si = season.loc[m_num]
            s_oil   = float(si['oil_s'])
            s_water = float(si['water_s'])
            s_liq   = float(si['liq_s'])
            s_wc    = float(si['wc_s'])
        else:
            s_oil = s_water = s_liq = s_wc = 1.0

        # decay للاتجاه (يتلاشى مع الزمن)
        decay = 0.92 ** i

        # تذبذب نسبي (% من القيمة الحالية)
        n_oil   = rng.normal(0, abs(cur_oil)   * vol_oil   * 0.9)
        n_water = rng.normal(0, abs(cur_water) * vol_water * 0.9) if cur_water > 1 else 0.0
        n_liq   = rng.normal(0, abs(cur_liq)   * vol_liq   * 0.9)
        n_wc    = rng.normal(0, abs(cur_wc)    * vol_wc    * 0.7)

        # تحديث بالاتجاه + تذبذب
        cur_oil   = cur_oil   + trend_oil   * decay + n_oil
        cur_water = cur_water + trend_water * decay + n_water
        cur_liq   = cur_liq   + trend_liq   * decay + n_liq
        cur_wc    = cur_wc    + trend_wc    * decay + n_wc

        # تطبيق الموسمية بخفة (قوة 30%)
        t_oil   = cur_oil   * (1 + (s_oil   - 1) * 0.3)
        t_water = cur_water * (1 + (s_water - 1) * 0.3)
        t_liq   = cur_liq   * (1 + (s_liq   - 1) * 0.3)
        t_wc    = cur_wc    * (1 + (s_wc    - 1) * 0.3)

        # ── تطبيق الحدود ──
        t_oil   = float(np.clip(t_oil,   floor_oil,   hist_max_oil))
        t_water = float(np.clip(t_water, floor_water, hist_max_water))
        t_liq   = float(np.clip(t_liq,   floor_liq,   hist_max_liq))
        t_wc    = float(np.clip(t_wc,    0.0, 100.0))

        # تناسق: نفط+ماء ≤ سائل كلي
        if t_oil + t_water > t_liq:
            t_liq = (t_oil + t_water) * 1.05

        # تحديث cur بالقيم المقيَّدة للحفاظ على الاستمرارية
        cur_oil   = t_oil
        cur_water = t_water
        cur_liq   = t_liq
        cur_wc    = t_wc

        # Random Forest للـ water cut
        if rf_model is not None and rf_scaler is not None:
            try:
                tgt   = date_cls(fcast.year, fcast.month, 15)
                feats = np.array([[t_liq, t_oil, t_water, float(tgt.toordinal())]])
                wc_rf = float(rf_model.predict(rf_scaler.transform(feats))[0])
                wc_rf = float(np.clip(wc_rf, 0.0, 100.0))
                calc  = (t_water / t_liq * 100) if t_liq > 0 else t_wc
                if abs(wc_rf - calc) > 20:
                    wc_rf = 0.3 * wc_rf + 0.7 * calc
                t_wc = wc_rf
            except Exception:
                pass

        gor_vals = [d['gor'] for d in all_data_sorted[-90:]
                    if d.get('gor') is not None and d['gor'] > 0]
        gor_val  = round(float(np.mean(gor_vals)), 2) if gor_vals else 0.0

        forecast_list.append({
            'date':         f"{fcast.year}-{fcast.month:02d}",
            'oil':          round(t_oil,   2),
            'water':        round(t_water, 2),
            'water_cut':    round(t_wc,    2),
            'gor':          gor_val,
            'total_liquid': round(t_liq,   2),
        })

    return forecast_list

    # ── بناء DataFrame شهري من البيانات التاريخية ──
    df_raw = pd.DataFrame(all_data_sorted)
    df_raw['date'] = pd.to_datetime(df_raw['date'])
    df_raw['ym']   = df_raw['date'].dt.to_period('M')

    monthly = df_raw.groupby('ym').agg(
        oil=('oil_production',    'mean'),
        water=('water_production', 'mean'),
        liq=('total_liquid',      'mean'),
        wc=('water_cut',          'mean'),
    ).reset_index()
    monthly['date'] = monthly['ym'].dt.start_time
    monthly = monthly.sort_values('date').reset_index(drop=True)

    n_hist = len(monthly)
    if n_hist < 3:
        return []

    # ── 1. نقطة البداية: آخر قيمة تاريخية فعلية ──
    last_oil   = float(monthly['oil'].iloc[-1])
    last_water = float(monthly['water'].iloc[-1])
    last_liq   = float(monthly['liq'].iloc[-1])
    last_wc    = float(monthly['wc'].iloc[-1])

    # ── 2. الاتجاه الشهري من آخر 6-12 شهراً (trend per month) ──
    window = min(12, n_hist)
    recent = monthly.tail(window).reset_index(drop=True)

    def monthly_trend(series):
        """حساب معدل التغير الشهري المتوسط (slope)"""
        y = series.values.astype(float)
        x = np.arange(len(y)).reshape(-1, 1)
        lr = LinearRegression()
        lr.fit(x, y)
        return float(lr.coef_[0])   # تغير لكل شهر

    trend_oil   = monthly_trend(recent['oil'])
    trend_water = monthly_trend(recent['water'])
    trend_liq   = monthly_trend(recent['liq'])
    trend_wc    = monthly_trend(recent['wc'])

    # ── 3. التذبذب الواقعي: بناءً على الفروق الشهرية الفعلية ──
    # نحسب الفرق بين كل شهر والسابق في البيانات التاريخية
    diff_oil   = monthly['oil'].diff().dropna().values
    diff_water = monthly['water'].diff().dropna().values
    diff_liq   = monthly['liq'].diff().dropna().values
    diff_wc    = monthly['wc'].diff().dropna().values

    # الانحراف المعياري للفروق = مقياس التذبذب الطبيعي
    vol_oil   = float(np.std(diff_oil))   if len(diff_oil)   > 1 else last_oil   * 0.03
    vol_water = float(np.std(diff_water)) if len(diff_water) > 1 else last_water * 0.03
    vol_liq   = float(np.std(diff_liq))   if len(diff_liq)   > 1 else last_liq   * 0.03
    vol_wc    = float(np.std(diff_wc))    if len(diff_wc)    > 1 else last_wc    * 0.03

    # ── 4. الموسمية: نسبة كل شهر نسبةً للمتوسط العام ──
    monthly['month_num'] = monthly['date'].dt.month
    season = monthly.groupby('month_num').agg(
        oil_s=('oil',   'mean'),
        water_s=('water','mean'),
        liq_s=('liq',   'mean'),
        wc_s=('wc',     'mean'),
    )
    mean_oil_all   = float(monthly['oil'].mean())   or 1.0
    mean_water_all = float(monthly['water'].mean()) or 1.0
    mean_liq_all   = float(monthly['liq'].mean())   or 1.0
    mean_wc_all    = float(monthly['wc'].mean())    or 1.0

    season['oil_s']   = (season['oil_s']   / mean_oil_all).clip(0.5, 1.5)
    season['water_s'] = (season['water_s'] / mean_water_all).clip(0.5, 1.5)
    season['liq_s']   = (season['liq_s']   / mean_liq_all).clip(0.5, 1.5)
    season['wc_s']    = (season['wc_s']    / mean_wc_all).clip(0.7, 1.3)
    season = season.fillna(1.0)

    # ── حدود تاريخية للتحقق من المنطقية ──
    hist_min_oil   = max(0.0, float(monthly['oil'].min())   * 0.5)
    hist_max_oil   = float(monthly['oil'].max())   * 1.5
    hist_min_water = max(0.0, float(monthly['water'].min()) * 0.5)
    hist_max_water = float(monthly['water'].max()) * 1.5
    hist_min_liq   = max(0.0, float(monthly['liq'].min())   * 0.5)
    hist_max_liq   = float(monthly['liq'].max())   * 1.5

    # بذرة عشوائية ثابتة للتكرارية
    rng = np.random.default_rng(seed=7)

    n_months = years_ahead * 12
    last_period = monthly['ym'].iloc[-1]
    forecast_list = []

    # القيم المتراكمة (تبدأ من آخر نقطة تاريخية)
    cur_oil   = last_oil
    cur_water = last_water
    cur_liq   = last_liq
    cur_wc    = last_wc

    for i in range(1, n_months + 1):
        fcast = last_period + i
        m_num = fcast.month

        # معامل الموسمية
        if m_num in season.index:
            si = season.loc[m_num]
            s_oil, s_water, s_liq, s_wc = (
                float(si['oil_s']), float(si['water_s']),
                float(si['liq_s']), float(si['wc_s'])
            )
        else:
            s_oil = s_water = s_liq = s_wc = 1.0

        # تطبيق الاتجاه + تذبذب عشوائي (Random Walk مع اتجاه)
        # الاتجاه يُخفَّف تدريجياً كلما ابتعدنا (decay) لمنع الانفجار
        decay = 0.97 ** i   # بعد 24 شهراً → 0.97^24 ≈ 0.48 تخميد للاتجاه

        noise_oil   = float(rng.normal(0, vol_oil   * 0.7))
        noise_water = float(rng.normal(0, vol_water * 0.7))
        noise_liq   = float(rng.normal(0, vol_liq   * 0.7))
        noise_wc    = float(rng.normal(0, vol_wc    * 0.5))

        cur_oil   = cur_oil   + trend_oil   * decay + noise_oil
        cur_water = cur_water + trend_water * decay + noise_water
        cur_liq   = cur_liq   + trend_liq   * decay + noise_liq
        cur_wc    = cur_wc    + trend_wc    * decay + noise_wc

        # تطبيق الموسمية (كنسبة مئوية طفيفة فوق الاتجاه)
        seasonal_strength = 0.4   # 0=بدون موسمية، 1=موسمية كاملة
        t_oil   = cur_oil   * (1 + (s_oil   - 1) * seasonal_strength)
        t_water = cur_water * (1 + (s_water - 1) * seasonal_strength)
        t_liq   = cur_liq   * (1 + (s_liq   - 1) * seasonal_strength)
        t_wc    = cur_wc    * (1 + (s_wc    - 1) * seasonal_strength)

        # ── حدود المنطقية ──
        t_oil   = float(np.clip(t_oil,   hist_min_oil,   hist_max_oil))
        t_water = float(np.clip(t_water, hist_min_water, hist_max_water))
        t_liq   = float(np.clip(t_liq,   hist_min_liq,   hist_max_liq))
        t_wc    = float(np.clip(t_wc,    0.0, 100.0))

        # تناسق: نفط+ماء ≤ سائل كلي
        if t_oil + t_water > t_liq:
            t_liq = (t_oil + t_water) * 1.05

        # Random Forest للـ water cut
        if rf_model is not None and rf_scaler is not None:
            try:
                tgt   = date_cls(fcast.year, fcast.month, 15)
                feats = np.array([[t_liq, t_oil, t_water, float(tgt.toordinal())]])
                wc_rf = float(rf_model.predict(rf_scaler.transform(feats))[0])
                wc_rf = float(np.clip(wc_rf, 0.0, 100.0))
                calc  = (t_water / t_liq * 100) if t_liq > 0 else t_wc
                if abs(wc_rf - calc) > 20:
                    wc_rf = 0.3 * wc_rf + 0.7 * calc
                t_wc = wc_rf
            except Exception:
                pass

        # تحديث القيم المتراكمة بالقيم المقيَّدة
        cur_oil   = t_oil
        cur_water = t_water
        cur_liq   = t_liq
        cur_wc    = t_wc

        # GOR من آخر 90 يوم تاريخي
        gor_vals = [d['gor'] for d in all_data_sorted[-90:]
                    if d.get('gor') is not None and d['gor'] > 0]
        gor_val  = float(np.mean(gor_vals)) if gor_vals else 0.0

        forecast_list.append({
            'date':         f"{fcast.year}-{fcast.month:02d}",
            'oil':          round(t_oil,   2),
            'water':        round(t_water, 2),
            'water_cut':    round(t_wc,    2),
            'gor':          round(gor_val, 2),
            'total_liquid': round(t_liq,   2),
        })

    return forecast_list


def predict_production_values_advanced(target_date, historical_data):
    """
    التنبؤ بقيم الإنتاج (Total Liquid, Oil, Water)
    باستخدام Polynomial Regression + Linear Regression + EWMA مع التحقق الشامل من المنطقية
    """
    if len(historical_data) < 2:
        if len(historical_data) == 1:
            return {
                'total_liquid':     float(historical_data[0]['total_liquid']),
                'oil_production':   float(historical_data[0]['oil_production']),
                'water_production': float(historical_data[0]['water_production'])
            }
        return None

    # بناء DataFrame وتنظيف البيانات
    rows = []
    for d in historical_data:
        try:
            rows.append({
                'date': datetime.strptime(d['date'], '%Y-%m-%d').date().toordinal(),
                'total_liquid':     float(d['total_liquid'])     if d['total_liquid']     is not None else 0.0,
                'oil_production':   float(d['oil_production'])   if d['oil_production']   is not None else 0.0,
                'water_production': float(d['water_production']) if d['water_production'] is not None else 0.0,
            })
        except Exception:
            continue

    if not rows:
        return None

    df = pd.DataFrame(rows).sort_values('date').reset_index(drop=True)

    # إذا كانت جميع القيم صفراً، ارجع لآخر قيمة غير صفرية
    if df['total_liquid'].sum() == 0 or df['oil_production'].sum() == 0:
        last_nz = df[df['total_liquid'] > 0].tail(1)
        if len(last_nz) > 0:
            return {
                'total_liquid':     float(last_nz['total_liquid'].iloc[0]),
                'oil_production':   float(last_nz['oil_production'].iloc[0]),
                'water_production': float(last_nz['water_production'].iloc[0])
            }
        return {
            'total_liquid':     max(float(df['total_liquid'].mean()), 1000.0),
            'oil_production':   max(float(df['oil_production'].mean()), 700.0),
            'water_production': max(float(df['water_production'].mean()), 300.0)
        }

    target_ordinal = target_date.toordinal() if hasattr(target_date, 'toordinal') else target_date.date().toordinal()
    min_date = df['date'].min()
    df['days_from_start'] = df['date'] - min_date
    target_days = float(target_ordinal - min_date)

    predictions = {}

    for col in ['total_liquid', 'oil_production', 'water_production']:
        values = df[col].values.astype(float)
        days   = df['days_from_start'].values.astype(float)

        # ── 1. Polynomial Regression (درجة 2 أو 3) ──
        poly_pred = None
        if len(df) >= 4:
            try:
                degree = min(3, max(2, len(df) // 3))
                poly_reg = Pipeline([
                    ('poly', PolynomialFeatures(degree=degree)),
                    ('linear', LinearRegression())
                ])
                poly_reg.fit(days.reshape(-1, 1), values)
                poly_pred = float(poly_reg.predict([[target_days]])[0])
                if np.isnan(poly_pred) or np.isinf(poly_pred):
                    poly_pred = None
            except Exception:
                poly_pred = None

        # ── 2. Linear Regression ──
        lr_pred = float(values.mean())
        if len(df) >= 2:
            try:
                lr = LinearRegression()
                lr.fit(days.reshape(-1, 1), values)
                lr_pred = float(lr.predict([[target_days]])[0])
                if np.isnan(lr_pred) or np.isinf(lr_pred):
                    lr_pred = float(values.mean())
            except Exception:
                lr_pred = float(values.mean())

        # ── 3. EWMA مع اتجاه (Trend) ──
        ewma_pred = float(values.mean())
        if len(df) >= 3:
            try:
                recent_count = max(3, len(df) // 2)
                recent_data  = df.tail(recent_count)
                recent_vals  = recent_data[col].values.astype(float)
                recent_days  = recent_data['days_from_start'].values.astype(float)

                alpha   = 0.3
                weights = np.array([alpha * (1 - alpha) ** (len(recent_vals) - 1 - i)
                                     for i in range(len(recent_vals))])
                weights /= weights.sum()
                ewma = float(np.average(recent_vals, weights=weights))

                if len(recent_days) >= 2:
                    day_diffs = np.diff(recent_days)
                    val_diffs = np.diff(recent_vals)
                    avg_daily_change = float(np.mean(val_diffs / day_diffs)) if np.sum(day_diffs) > 0 else 0.0
                    days_ahead = target_days - recent_days[-1]
                    ewma_pred  = ewma + avg_daily_change * days_ahead
                    if np.isnan(ewma_pred) or np.isinf(ewma_pred):
                        ewma_pred = ewma
                else:
                    ewma_pred = ewma
            except Exception:
                ewma_pred = float(values.mean())

        # ── دمج التنبؤات الثلاثة ──
        if poly_pred is not None:
            final_pred = 0.5 * poly_pred + 0.3 * lr_pred + 0.2 * ewma_pred
        else:
            final_pred = 0.6 * lr_pred + 0.4 * ewma_pred

        # إذا كانت النتيجة صفراً أو سالبة أو NaN → ارجع للمتوسط
        if final_pred <= 0 or np.isnan(final_pred) or np.isinf(final_pred):
            final_pred = float(values.mean())

        predictions[col] = max(0.01, final_pred)

    # ── التحقق من المنطقية مقارنةً بالبيانات التاريخية ──
    for col in ['total_liquid', 'oil_production', 'water_production']:
        max_val  = float(df[col].max())
        min_val  = float(df[col].min())
        mean_val = float(df[col].mean())

        # لا تسمح بأكثر من 3 أضعاف القيمة القصوى
        if predictions[col] > max_val * 3:
            predictions[col] = mean_val * 1.2

        # لا تسمح بأقل من 10% من القيمة الدنيا (غير الصفرية)
        elif min_val > 0 and predictions[col] < min_val * 0.1:
            predictions[col] = mean_val * 0.8

    # ── التحقق من التناسق الداخلي (نفط + ماء ≤ سائل كلي) ──
    total_sum = predictions['oil_production'] + predictions['water_production']

    if predictions['total_liquid'] < total_sum:
        predictions['total_liquid'] = total_sum * 1.05

    if total_sum > predictions['total_liquid'] * 1.15:
        ratio = (predictions['total_liquid'] * 0.95) / total_sum
        predictions['oil_production']   *= ratio
        predictions['water_production'] *= ratio

    if predictions['oil_production'] > predictions['total_liquid'] * 0.99:
        predictions['oil_production'] = predictions['total_liquid'] * 0.95

    if predictions['water_production'] > predictions['total_liquid'] * 0.99:
        predictions['water_production'] = predictions['total_liquid'] * 0.95

    return predictions

def train_random_forest_model():
    """تدريب Random Forest للتنبؤ بنسبة الماء"""
    data = get_all_data()
    if len(data) < 5:
        return None, None
    df = pd.DataFrame([{
        'total_liquid': float(d['total_liquid']),
        'oil_production': float(d['oil_production']),
        'water_production': float(d['water_production']),
        'date_ordinal': datetime.strptime(d['date'], '%Y-%m-%d').date().toordinal(),
        'water_cut': float(d['water_cut'])
    } for d in data])
    X = df[['total_liquid', 'oil_production', 'water_production', 'date_ordinal']]
    y = df['water_cut'].values
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    model = RandomForestRegressor(n_estimators=100, random_state=42, max_depth=10)
    model.fit(X_scaled, y)
    return model, scaler

def predict_water_cut_with_rf(target_date, predicted_values, rf_model, rf_scaler):
    """التنبؤ بنسبة الماء مع التحقق المزدوج (Random Forest + نسبة محسوبة)"""
    calculated = (
        (predicted_values['water_production'] / predicted_values['total_liquid'] * 100)
        if predicted_values['total_liquid'] > 0 else 0.0
    )

    if rf_model is None or rf_scaler is None:
        return calculated

    try:
        date_ordinal = target_date.toordinal() if hasattr(target_date, 'toordinal') else target_date.date().toordinal()
        features = np.array([[
            predicted_values['total_liquid'],
            predicted_values['oil_production'],
            predicted_values['water_production'],
            float(date_ordinal)
        ]])
        features_scaled = rf_scaler.transform(features)
        pred = float(rf_model.predict(features_scaled)[0])

        if np.isnan(pred) or np.isinf(pred):
            return calculated

        pred = max(0.0, min(100.0, pred))

        # إذا كان الفرق بين RF والنسبة المحسوبة أكثر من 20%، استخدم المتوسط المرجح
        if abs(pred - calculated) > 20:
            pred = 0.3 * pred + 0.7 * calculated

        return pred

    except Exception:
        return calculated

def get_monthly_aggregates(df):
    """تجميع البيانات اليومية إلى شهرية"""
    df['date'] = pd.to_datetime(df['date'])
    df['year_month'] = df['date'].dt.to_period('M')
    monthly = df.groupby('year_month').agg({
        'oil_production': 'mean',
        'water_production': 'mean',
        'water_cut': 'mean',
        'gor': 'mean',
        'total_liquid': 'mean'
    }).reset_index()
    monthly['date'] = monthly['year_month'].dt.start_time
    monthly = monthly.sort_values('date')
    return monthly

@app.route('/api/forecast', methods=['GET'])
def api_forecast():
    years_ahead = request.args.get('years_ahead', default=2, type=int)
    if years_ahead < 0:
        years_ahead = 0

    all_data = get_all_data()
    if len(all_data) < 3:
        return jsonify({'historical': [], 'forecast': [], 'stats': None})

    all_data_sorted = sorted(all_data, key=lambda x: x['date'])

    # ── البيانات التاريخية الشهرية ──
    df_hist = pd.DataFrame(all_data_sorted)
    monthly_hist = get_monthly_aggregates(df_hist)
    historical = []
    for _, row in monthly_hist.iterrows():
        historical.append({
            'date':         row['date'].strftime('%Y-%m'),
            'oil':          round(float(row['oil_production']), 2),
            'water':        round(float(row['water_production']), 2),
            'water_cut':    round(float(row['water_cut']), 2),
            'gor':          round(float(row['gor']), 2) if not pd.isna(row['gor']) else 0,
            'total_liquid': round(float(row['total_liquid']), 2)
        })

    # ── التنبؤ المستقبلي (سريع — شهري مباشرة) ──
    forecast = []
    if years_ahead > 0:
        rf_model, rf_scaler = train_random_forest_model()
        forecast = build_forecast_monthly(all_data_sorted, years_ahead, rf_model, rf_scaler)

    stats = {
        'method': 'Polynomial Trend + Monthly Seasonality + Statistical Noise + Random Forest (Water Cut)',
        'features': 'Oil, Water, Total Liquid, Water Cut — Monthly Aggregates',
        'training_samples': len(all_data_sorted)
    }

    return jsonify({'historical': historical, 'forecast': forecast, 'stats': stats})

# ------------------- STARTUP -------------------
def _background_train():
    """تدريب الموديل في الخلفية بدون تأخير فتح السيرفر"""
    df = db_to_dataframe()
    if len(df) >= 10:
        print("🧠 Training neural network in background...")
        try:
            model_instance.train(df)
            save_model()
            print(f"✅ Training completed (R²={model_instance.score:.3f})")
        except Exception as e:
            print(f"⚠️ Training failed: {e}")

def startup():
    init_db()
    load_model()  # ← حمّل الموديل المحفوظ أولاً (سريع جداً)

    conn = sqlite3.connect(app.config['DATABASE'])
    count = conn.execute("SELECT COUNT(*) FROM production_data").fetchone()[0]
    conn.close()

    # تحميل Excel مرة واحدة فقط إذا كانت قاعدة البيانات فارغة
    if count == 0:
        excel_file = os.path.join(BASE_DIR, 'تحسين_الانتاج.xlsx')
        if os.path.exists(excel_file):
            print("📥 Loading data from Excel file...")
            try:
                df = parse_excel(excel_file)
                conn2 = sqlite3.connect(app.config['DATABASE'])
                added = 0
                for _, row in df.iterrows():
                    vals = {c: (None if pd.isna(row.get(c, np.nan)) else row.get(c)) for c in
                            ['date','total_liquid','oil_production','water_production','water_cut','upstream_pres','downstream_pres','choke_size','gor']}
                    if vals.get('oil_production') is None: continue
                    conn2.execute("""INSERT INTO production_data 
                        (date, total_liquid, oil_production, water_production, water_cut, upstream_pres, downstream_pres, choke_size, gor)
                        VALUES (:date,:total_liquid,:oil_production,:water_production,:water_cut,:upstream_pres,:downstream_pres,:choke_size,:gor)""", vals)
                    added += 1
                conn2.commit()
                conn2.close()
                print(f"✅ Successfully added {added} records")
            except Exception as e:
                print(f"⚠️ Error loading Excel: {e}")

    # تدريب الموديل في الخلفية إذا لم يكن محفوظاً — السيرفر يفتح فوراً
    if not model_instance.is_trained:
        t = threading.Thread(target=_background_train, daemon=True)
        t.start()

if __name__ == '__main__':
    startup()
    print("\n🚀 Server running on http://127.0.0.1:5000\n")
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)